'''
############################################################################################################################
Risk Assessment for Generic Cyber-Physical Systems (CPS) using Bayesian Belief Networks (BBN) based on AutomationML Models
(Based on code by Pushparaj BHOSALE: https://github.com/Pbhosale1991/AML-BBN-RA)
Author: Huang Shaofei
Last update: 2025-01-18                                                         
Latest version at https://github.com/shaofeihuang/CPS-Risk_Assessment-BBN
###################################################################################
'''
import numpy as np
import xml.etree.ElementTree as ET
import networkx as nx
import matplotlib.pyplot as plt
import itertools
import openpyxl
import math
import random
from datetime import datetime
from collections import defaultdict
from openpyxl.styles import PatternFill
from pgmpy.models import BayesianNetwork
from pgmpy.factors.discrete import TabularCPD
from pgmpy.inference import VariableElimination

############################
# Section 1: Program Start #
############################
instanceHierarchyTag=".//{http://www.dke.de/CAEX}InstanceHierarchy"
internalElementTag=".//{http://www.dke.de/CAEX}InternalElement"
externalInterfaceTag=".//{http://www.dke.de/CAEX}ExternalInterface"
AttributeTag=".//{http://www.dke.de/CAEX}Attribute"
ValueTag=".//{http://www.dke.de/CAEX}Value"
internalLinkTag=".//{http://www.dke.de/CAEX}InternalLink"

if __name__ == "__main__":
    amlFile = ET.parse('Generic_CPS_Mitigation.aml')
    root = amlFile.getroot()

def get_valid_date():
    while True:
        date_input = input("Enter system installation date (in the format YYYY-MM-DD) or leave blank for default (2024-01-01): ")
        if not date_input:
            return "2024-01-01"  # Default date if no input
        try:
            datetime.strptime(date_input, "%Y-%m-%d")
            return date_input
        except ValueError:
            print("Invalid date format. Please try again.")
  
def calculate_days_and_hours(start_date):
    start_date = datetime.strptime(start_date, "%Y-%m-%d")
    reference_date_input = input("Enter reference date (YYYY-MM-DD) or leave blank for current date: ")
    if reference_date_input.strip():
        reference_date = datetime.strptime(reference_date_input, "%Y-%m-%d")
    else:
        reference_date = datetime.now()
    print("[*]: Reference date:", reference_date)
    time_difference = reference_date - start_date
    days = time_difference.days
    remaining_seconds = time_difference.seconds
    remaining_hours = remaining_seconds // 3600
    return days, remaining_hours

start_date_str = get_valid_date()
print("[*]: System installation date:", start_date_str)
days, hours = calculate_days_and_hours(start_date_str)
t = days * 4 + (24 - hours)
print("Time since installation:", days, "days and ", hours, "hours (Total:", t, "hours)\n")

sap_input = input("Enter probability of successful security attack (SA) (0.01% < value < 10%) or leave blank for default (1%): ")

if not sap_input:
    sap_percent = 1  # Default value if no input
else:
    sap_percent = float(sap_input)

if 0.01 <= sap_percent <= 10:
    sap = sap_percent / 100
else:
    print("ERROR: Input a valid SA value.")
############################
#     End of Section 1     #
############################

#############################################
# Section 2: AML Model Attribute Extraction #
#############################################
allinone_attrib = []  
allinone_tags = []
allinone_text = []
name_id_tag_list = []
name_list = []
id_list = []
tag_list = []
RefPartnerBegin_list = []
RefPartnerTerminate_list = []
InternalLinks = []
probability_data = []
HazardinSystem = []
VulnerabilityinSystem = []
external_interfaces_list = []
connections = []
interface_to_element_map = {}
connections_mapped = []
result_list = []
number_of_dependents = []
AutomationPyramidStatus = {
    'AssetOfICS/Hardware/Machine': [],
    'AssetOfICS/Hardware/Process device/Actuator': [],
    'AssetOfICS/Hardware/Process device/Sensor': [],
    'AssetOfICS/Hardware/Process device/Controller': [],
    'AssetOfICS/Hardware/Process device/Workstation': [],
    'AssetOfICS/User': []
}
max_num_parents = 0
max_num_children = 0

for k in root.findall('.//'):
    allinone_attrib.append(k.attrib)
    allinone_tags.append(k.tag)
    allinone_text.append(k.text)

for i, component_attrib in enumerate(allinone_attrib):
    name = component_attrib.get('Name')
    ID = component_attrib.get('ID')
    RPA = component_attrib.get('RefPartnerSideA')
    RPB = component_attrib.get('RefPartnerSideB')
    if name:
        name_list.append(name)
    if ID:
        id_list.append(ID)
    if name and ID:
        tag = allinone_tags[i]
        tag_list.append(tag)
        name_id_tag_list.append({'Name': name, 'ID': ID, 'Tag': tag})
    if RPA:
        RefPartnerBegin_list.append(RPA)
    if RPB:
        RefPartnerTerminate_list.append(RPB)
    if RPA and RPB:
        #InternalLinks.append({RPA, RPB, tag})
        InternalLinks.append({RPA, RPB})

internal_elements = root.findall(internalElementTag)

def get_attribute_value(internal_element, attribute_name):
    attribute_tag = internal_element.find(f".//{{http://www.dke.de/CAEX}}Attribute[@Name='{attribute_name}']")
    if attribute_tag is not None:
        value_element = attribute_tag.find(ValueTag)
        if value_element is not None:
            return float(value_element.text)
    return None

def calculate_probability_of_failure(failure_rate_value, t):
    failure_rate = float(failure_rate_value)
    return 1 - math.exp(-(failure_rate * t))

def calculate_probability_of_human_error(human_error_percentage_value, t):
    human_error_in_percent = float(human_error_percentage_value)
    human_error_rate = human_error_in_percent / (100 * 8760)
    return 1 - math.exp(-(human_error_rate * t))

def generate_cpd_values_hazard(num_parents):
    cpd_values = [[0] * (2 ** num_parents) for _ in range(2)]
    for i in range(2 ** num_parents):
        num_ones = bin(i).count('1')
        cpd_values[0][i] = (num_parents - num_ones) / num_parents
        cpd_values[1][i] = 1 - cpd_values[0][i]
    return cpd_values

def generate_cpd_values_impact(num_parents):
    cpd_values = [[0] * (2 ** num_parents) for _ in range(2)]
    current_entry = next((entry for entry in result_list if entry['Element'] == node), None)
    for i in range(2 ** num_parents):
        num_ones = bin(i).count('1')
        cpd_values[0][i] = (num_parents - num_ones) / num_parents
        cpd_values[1][i] = 1 - cpd_values[0][i]
        if num_parents - num_ones == 0:
            cpd_values[0][i] = current_entry['Number of children']
            cpd_values[1][i] = 1 - cpd_values[0][i]
    return cpd_values

for internal_element in internal_elements:
    internal_element_id = internal_element.get('ID')
    internal_element_name = internal_element.get('Name')
    ref_base_system_unit_path = internal_element.get('RefBaseSystemUnitPath')

    failure_rate_value = get_attribute_value(internal_element, 'FailureRatePerHour')
    probability_of_failure = None
    if failure_rate_value is not None:
        probability_of_failure = calculate_probability_of_failure(failure_rate_value, t)
    else:
        probability_of_failure = 0
    
    probability_of_exposure_value = get_attribute_value(internal_element, 'Probability of Exposure')
    probability_of_exposure = None
    if probability_of_exposure_value is not None:
        probability_of_exposure = probability_of_exposure_value
    else:
        probability_of_exposure = 0
    
    probability_of_impact_value = get_attribute_value(internal_element, 'Probability of Impact')
    probability_of_impact_vulnerability = None
    if probability_of_impact_value is not None:
        probability_of_impact_vulnerability = probability_of_impact_value
    else:
        probability_of_impact_vulnerability = 0

    human_error_percentage_value = get_attribute_value(internal_element, 'HumanErrorEstimationPercentage')
    probability_of_human_error = None
    if human_error_percentage_value is not None:
        probability_of_human_error = calculate_probability_of_human_error(human_error_percentage_value, t)
    else:
        probability_of_human_error = 0

    internal_element_data = {
        'ID': internal_element_id,
        'Name': internal_element_name,
        'Probability of Failure': probability_of_failure,
        'Probability of Exposure': probability_of_exposure,
        'Probability of Impact' : probability_of_impact_vulnerability,
        'Probability of Human Error': probability_of_human_error,
        'RefBaseSystemUnitPath': ref_base_system_unit_path
    }
    
    if ref_base_system_unit_path in AutomationPyramidStatus:
        AutomationPyramidStatus[ref_base_system_unit_path].append(internal_element_data)
    elif ref_base_system_unit_path == 'HazardforSystem/Hazard':
        HazardinSystem.append(internal_element_data)
    elif ref_base_system_unit_path == 'VulnerabilityforSystem/Vulnerability':
        VulnerabilityinSystem.append(internal_element_data)
    
    probability_data.append(internal_element_data)

## Data Check
#for data in probability_data:
#    print("ID:", data['ID'], "Name:", data['Name'], "RefSystemUnitPath:", data['RefBaseSystemUnitPath'],
#          "Prob of Failure:", data['Probability of Failure'], "Prob of Exposure:", data['Probability of Exposure'], "Prob of Human Error:", data['Probability of Human Error'])

for internal_element in internal_elements:
    external_interfaces = internal_element.findall(externalInterfaceTag)
    if len(external_interfaces) < 5: # there are 5 external interfaces links Network, process, User, hazard, and vulnerability based
        internal_element_id = internal_element.get('ID')
        internal_element_name = internal_element.get('Name')
        for external_interface in external_interfaces:
            external_interface_id = external_interface.get('ID')
            external_interface_name = external_interface.get('Name')
            external_interface_ref_base_class_path = external_interface.get('RefBaseClassPath')
            if external_interface_ref_base_class_path != 'ConnectionBetnAssets/Network based':
                external_interface_info = {
                    'InternalElement ID': internal_element_id,
                    'InternalElement Name': internal_element_name,
                    'ExternalInterface ID': external_interface_id,
                    'ExternalInterface Name': external_interface_name,
                    'ExternalInterface RefBaseClassPath': external_interface_ref_base_class_path
                    }            
                external_interfaces_list.append(external_interface_info)

for external_interface in external_interfaces_list:
    external_interface_id = external_interface['ExternalInterface ID']
    internal_element_id = external_interface['InternalElement ID']
    interface_to_element_map[external_interface_id] = internal_element_id

for internal_link in root.findall(internalLinkTag):
    ref_partner_a = internal_link.get('RefPartnerSideA')
    ref_partner_b = internal_link.get('RefPartnerSideB')
    if ref_partner_a in interface_to_element_map and ref_partner_b in interface_to_element_map:
        internal_element_a = interface_to_element_map[ref_partner_a]
        internal_element_b = interface_to_element_map[ref_partner_b]
        connection = {'from': internal_element_a, 'to': internal_element_b}
        connections.append(connection)

for connection in connections:
    from_interface = connection['from']
    to_interface = connection['to']
    if from_interface in interface_to_element_map:
        from_element = interface_to_element_map[from_interface]
    else:
        from_element = from_interface  # If from_interface is not in the map, assume it's already an internal element ID
    if to_interface in interface_to_element_map:
        to_element = interface_to_element_map[to_interface]
    else:
        to_element = to_interface  # If to_interface is not in the map, assume it's already an internal element ID
    mapped_connection = {'from': from_element, 'to': to_element}
    connections_mapped.append(mapped_connection)

connections_from_to = defaultdict(list)
connections_to_from = defaultdict(list)
total_elements = set()

for connection in connections_mapped:
    from_element = connection['from']
    to_element = connection['to']
    total_elements.add(from_element)
    total_elements.add(to_element)
    connections_from_to[from_element].append(to_element)
    connections_to_from[to_element].append(from_element)

connections_result_FT = [{'from': k, 'to': v} for k, v in connections_from_to.items()]
connections_result_TF = [{'from': v, 'to': k} for k, v in connections_to_from.items()]
number_of_children =  [{'Element': k, 'Number of children': len(v)} for k, v in connections_from_to.items()]
number_of_parents =  [{'Element': k, 'Number of parents': len(v)} for k, v in connections_to_from.items()]

for element in total_elements:
    child = next((c for c in number_of_children if c['Element'] == element), {'Number of children': 0})
    parent = next((p for p in number_of_parents if p['Element'] == element), {'Number of parents': 0})
    total_dependents = child['Number of children'] + parent['Number of parents']
    result_dict = {
        'Element': element,
        'Number of children': child['Number of children'],
        'Number of parents': parent['Number of parents'],
        'Total Dependents': total_dependents
    }

    for key in result_dict:
        if isinstance(result_dict[key], (int, float)):
            result_dict[key] /= len(total_elements)
    
    result_list.append(result_dict)
    parent = next((p for p in number_of_parents if p['Element'] == element), {'Number of parents': 0})
    num_parents = parent['Number of parents']

    if num_parents > max_num_parents:
        max_num_parents = num_parents
    
    num_children = child['Number of children']

    if num_children > max_num_children:
        max_num_children = num_children

#############################################
#             End of Section 2              #
#############################################

############################################################################
# Section 3: Bayesian Belief Network (BBN) Implementation Helper Functions #
############################################################################
def generate_cpd_values_occur(num_states, num_parents, hazard_node=False, vulnerability_node=False, process_node=False):
    cpd_values = np.zeros((num_states, 2 ** num_parents))

    if hazard_node:
        if num_parents == 0:
            cpd_values[0, 0] = 0.5
            cpd_values[1, 0] = 0.5
        elif num_parents == 1:
            cpd_values[0, 0] = 1
            cpd_values[0, 1] = 0
            cpd_values[1, 0] = 1 - cpd_values[0, 0]
            cpd_values[1, 1] = 1 - cpd_values[0, 1]
        elif 2 <= num_parents <= max_num_parents:
            cpd_values=generate_cpd_values_hazard(num_parents)

    elif vulnerability_node:
        probability_of_exposure_for_node = matching_vulnerability_nodes[0]['Probability of Exposure']
        pofe = float(probability_of_exposure_for_node)        
        if num_parents == 0:
            cpd_values[0, 0] = pofe * sap
            cpd_values[1, 0] = 1 - pofe * sap
        elif num_parents >= 1:
            cpd_values[0, :-1] = pofe * sap  	# parent vulnerability is exposed
            cpd_values[1, :-1] = 1 - pofe * sap # parent vulnerability is unexposed
            cpd_values[0, -1] = 0	            # parent vulnerability is exposed
            cpd_values[1, -1] = 1
    
    elif process_node:
        ref_base_for_node = matching_process_nodes[0]['RefBaseSystemUnitPath']    
        if ref_base_for_node in ['AssetOfICS/Hardware/Process device/Actuator',
                                 'AssetOfICS/Hardware/Process device/Sensor',
                                 'AssetOfICS/Hardware/Machine', 
                                 'AssetOfICS/Hardware/Process device/Controller',
                                 'AssetOfICS/Hardware/Process device/Workstation']:
            probability_of_failure_for_node = matching_process_nodes[0]['Probability of Failure']
            if probability_of_failure_for_node:
                poff = float(probability_of_failure_for_node)
                cpd_values[0, :-1] = 1
                cpd_values[1, :-1] = 0
                cpd_values[0, -1] = poff
                cpd_values[1, -1] = 1 - poff
            else:
                cpd_values[0, :-1] = 1
                cpd_values[1, :-1] = 0
                cpd_values[0, -1] = 0
                cpd_values[1, -1] = 1
        elif ref_base_for_node == 'AssetOfICS/User':
            probability_of_human_error_for_node = matching_process_nodes[0]['Probability of Human Error']
            pofhe = float(probability_of_human_error_for_node)
            cpd_values[0, 0] = pofhe
            cpd_values[1, 0] = 1 - pofhe
        else:
            probability_of_failure_for_node = matching_process_nodes[0]['Probability of Failure']
            poff = float(probability_of_failure_for_node)
            cpd_values[0, 0] = poff
            cpd_values[1, 0] = 1 - poff

    cpd_values /= np.sum(cpd_values, axis=0)  # Normalize the CPD values
    return cpd_values.reshape((num_states, -1))

def generate_cpd_values_impact_(num_states, num_parents, hazard_node=False, vulnerability_node=False, process_node=False):
    cpd_values = np.zeros((num_states, 2 ** num_parents))
    current_entry = next((entry for entry in result_list if entry['Element'] == node), None)

    if hazard_node:
        if num_parents == 0:
            cpd_values[0, 0] = 0.5
            cpd_values[1, 0] = 0.5
        elif num_parents == 1:
            cpd_values[0, 0] = 1
            cpd_values[0, 1] = 0
            cpd_values[1, 0] = 1 - cpd_values[0, 0]
            cpd_values[1, 1] = 1 - cpd_values[0, 1]    
        elif 2 <= num_parents <= max_num_parents:
            cpd_values=generate_cpd_values_impact(num_parents)
    
    elif vulnerability_node:
        probability_of_impact_for_node = matching_vulnerability_nodes[0]['Probability of Impact']
        pofi = float(probability_of_impact_for_node)
        if num_parents == 0:
            cpd_values[0, 0] = pofi * sap
            cpd_values[1, 0] = 1 - pofi * sap
        elif num_parents >= 1:
            cpd_values[0, :-1] = 1  	        # parent vulnerability is exposed
            cpd_values[1, :-1] = 0              # parent vulnerability is unexposed
            cpd_values[0, -1] = pofi * sap	    # parent vulnerability is exposed
            cpd_values[1, -1] = 1 - pofi * sap	# parent vulnerability is unexposed
    
    elif process_node:
        ref_base_for_node = matching_process_nodes[0]['RefBaseSystemUnitPath']
        if ref_base_for_node in ['AssetOfICS/Hardware/Process device/Actuator',
                                 'AssetOfICS/Hardware/Process device/Sensor',
                                 'AssetOfICS/Hardware/Machine', 
                                 'AssetOfICS/Hardware/Process device/Controller',
                                 'AssetOfICS/Hardware/Process device/Workstation']:
            probability_of_failure_for_node = matching_process_nodes[0]['Probability of Failure']
            if probability_of_failure_for_node:
                cpd_values[0, :-1] = 1
                cpd_values[1, :-1] = 0
                cpd_values[0, -1] = current_entry['Number of children']
                cpd_values[1, -1] = 1 - current_entry['Number of children']
            else:
                cpd_values[0, :-1] = 1
                cpd_values[1, :-1] = 0
                cpd_values[0, -1] = 0
                cpd_values[1, -1] = 1   
        elif ref_base_for_node == 'AssetOfICS/User':
            cpd_values[0, 0] = current_entry['Number of children']
            cpd_values[1, 0] = 1 - current_entry['Number of children']
        else:
            cpd_values[0, 0] = current_entry['Number of children']
            cpd_values[1, 0] = 1 - current_entry['Number of children']

    cpd_values /= np.sum(cpd_values, axis=0)  # Normalize the CPD values
    return cpd_values.reshape((num_states, -1))

def shortest_path_length(graph, start_node, end_node):
    try:
        # Using networkx's shortest path length function
        length = nx.shortest_path_length(graph, source=start_node, target=end_node)
        return length
    except nx.NetworkXNoPath:
        # If no path exists between the nodes
        return float('inf')

def select_start_end_nodes(total_elements):
    start_node = random.choice(list(total_elements))
    end_node = random.choice(list(total_elements - {start_node}))  # Ensuring end_node is different from start_node
    return start_node, end_node

############################################################################
#                            End of Section 3                              #
############################################################################

#########################################################################
# Section 4: Bayesian Belief Network (BBN) Implementation for Occurrence #
#########################################################################
cpds = {}
cpd_values_list = []
nodes_and_numberofParents =[]
path_length_betn_nodes=[]
path_length_betn_nodes_final=[]
path_length_final_node = []

bbn = BayesianNetwork()
connections = connections_mapped
bbnNodes=bbn.add_nodes_from(total_elements)
bbn.add_edges_from([(connection['from'], connection['to']) for connection in connections])

for node in bbn.nodes():
    num_parents = len(bbn.get_parents(node))
    num_states = 2  # Assuming binary states for each node
    matching_hazard_nodes = [element for element in HazardinSystem if element['ID'] == node]
    matching_vulnerability_nodes = [element for element in VulnerabilityinSystem if element['ID'] == node]
    matching_process_nodes = [element for element in probability_data if element['ID'] == node]

    cpd_values = None

    if matching_hazard_nodes:
        hazard_node = True
        cpd_values = generate_cpd_values_occur(num_states, num_parents, hazard_node=True)
    elif matching_vulnerability_nodes:
        vulnerability_node = True
        cpd_values = generate_cpd_values_occur(num_states, num_parents, vulnerability_node=True)
    elif matching_process_nodes:
        process_node = True
        cpd_values = generate_cpd_values_occur(num_states, num_parents, process_node=True)

    cpd = TabularCPD(variable=node, variable_card=num_states, values=cpd_values,
                     evidence=bbn.get_parents(node), evidence_card=[2] * num_parents)

    cpds[node] = cpd
    cpd_values_list.append((node, cpd_values.tolist(), cpd.variables, cpd.cardinality))
    
bbn.add_cpds(*cpds.values())
bbn_graph = bbn.to_markov_model()

last_node = None
for element1 in HazardinSystem:
    node1=element1['ID']
    for element2 in result_list:
        node2=element2['Element']
        child_num = element2['Number of children']
        if node1 == node2:
            if child_num == 0:
                last_node = node1

print("\n[*] Last node in BBN (Occurrence):", last_node)

for node1, node2 in itertools.product(total_elements, repeat=2):
    if node1 == node2:
        path_length_betn_nodes.append((node1, node2, 0))
    else:
        path_length = shortest_path_length(bbn_graph, node1, node2)
        if path_length == float('inf'):
            path_length_betn_nodes.append((node1, node2, "No path"))
        else:
            path_length_betn_nodes_final.append((node1, node2, path_length, 1/path_length))
            path_length_betn_nodes.append({'Node1': node1, 'Node2': node2, 
                               'Number of hops': path_length, 
                               'Probability': 1/path_length})
            if node2==last_node:
                path_length_final_node.append((node1, last_node, path_length, 1/path_length))

#########################################################################
#                          End of Section 4                             #
#########################################################################

######################################################################
# Section 5: Bayesian Belief Network (BBN) Implementation for Impact #
######################################################################
bbn_impact=BayesianNetwork()
bbn_impact.add_edges_from([(connection['from'], connection['to']) for connection in connections])
cpds = {}
cpd_values_list_impact = []
nodes_and_numberofParents =[]

for node in bbn_impact.nodes():
    num_parents = len(bbn.get_parents(node))
    num_states = 2
    cpd_values = None

    matching_hazard_nodes = [element for element in HazardinSystem if element['ID'] == node]
    matching_vulnerability_nodes = [element for element in VulnerabilityinSystem if element['ID'] == node]
    matching_process_nodes = [element for element in probability_data if element['ID'] == node]

    if matching_hazard_nodes:
        hazard_node = True
        cpd_values = generate_cpd_values_impact_(num_states, num_parents, hazard_node=True)
    elif matching_vulnerability_nodes:
        vulnerability_node = True
        cpd_values = generate_cpd_values_impact_(num_states, num_parents, vulnerability_node=True)
    elif matching_process_nodes:
        process_node = True
        cpd_values = generate_cpd_values_impact_(num_states, num_parents, process_node=True)

    cpd = TabularCPD(variable=node, variable_card=num_states, values=cpd_values,
                     evidence=bbn.get_parents(node), evidence_card=[2] * num_parents)

    cpds[node] = cpd
    cpd_values_list.append((node, cpd_values.tolist(), cpd.variables, cpd.cardinality))

bbn_impact.add_cpds(*cpds.values())
######################################################################
#                          End of Section 5                          #
######################################################################

#############################
# Section 6: BBN Operations #
#############################
print("[*] Checking BBN (Occurrence) structure consistency:", bbn.check_model())
print("[*] Checking BBN (Severity/Impact) structure consistency:", bbn_impact.check_model())

inference = VariableElimination(bbn)
inference2 = VariableElimination(bbn_impact)

# Plot BBN visual
graph = nx.DiGraph()
graph.add_nodes_from(bbn.nodes())
graph.add_edges_from(bbn.edges())
pos = nx.kamada_kawai_layout(graph, scale=2)
nx.draw_networkx_nodes(graph, pos, node_color='lightblue', node_size=300)
nx.draw_networkx_edges(graph, pos, arrows=True, arrowstyle='->', arrowsize=10)
nx.draw_networkx_labels(graph, pos)
plt.title("Bayesian Belief Network")
plt.axis('off')
#plt.savefig('bbn_plot_shell_layout.png', format='png', dpi=300, bbox_inches='tight')
plt.show()

# Shortest paths
print('[-] Find shortest path in BBN (Occurrence)')
graph = nx.DiGraph(bbn.edges)

valid_nodes = {"v1", "v2", "v3", "v4", "v5", "v6"}

while True:
    source_node_input = input("Enter source node (V1-V6) or leave blank for default (V2: Tampering): ").strip()
    if not source_node_input:  # Default to "V2" if input is blank
        source_node = "V2"
        break
    # Convert input to lowercase and check if it's valid
    if source_node_input.lower() in valid_nodes:
        source_node = source_node_input.upper()  # Convert to uppercase for consistency
        break
    else:
        print("[!] Invalid input. Please enter a valid node (V1-V6).")

target_node = "CPS_Termination"

try:
    shortest_path = nx.shortest_path(graph, source=source_node, target=target_node)
    print(f"[*] Shortest path from {source_node} to {target_node} is: {shortest_path}")
except nx.NetworkXNoPath:
    print(f"[!] No path exists between {source_node} and {target_node}.")
except nx.NodeNotFound as e:
    print(f"[!] Error: {e}")

print('[-] Find shortest path in BBN (Severity/Impact)')
graph = nx.DiGraph(bbn_impact.edges)
try:
    shortest_path = nx.shortest_path(graph, source=source_node, target=target_node)
    print(f"[*] Shortest path from {source_node} to {target_node} is: {shortest_path}")
except nx.NetworkXNoPath:
    print(f"[!] No path exists between {source_node} and {target_node}.")
except nx.NodeNotFound as e:
    print(f"[!] Error: {e}")

# Compute risk score
risk_measurements = []
for node in total_elements:
    prob_node = inference.query(variables=[node])
    impact_node = inference2.query(variables=[node])
    for element in path_length_final_node:
        if node == element[0]:
            cpd_prob = prob_node.values
            cpd_imp = element[3]
            cpd_impact = impact_node.values        
            risk_v1=cpd_prob[0] * cpd_imp
            risk_v2=cpd_prob[0] * cpd_impact[0]
            #print(cpd_prob[0], cpd_imp, risk)
            risk_measurements.append([node, cpd_prob[0], cpd_imp, cpd_impact[0], risk_v1, risk_v2])

# Sort nodes according to prob
node_prob_dict = {}
for node in total_elements:
    if node == last_node:
        pass
    else:
        prob_P3L_if_node_fail = inference.query(variables=[last_node], evidence={node:0})
        prob_P3L_if_node_fail_values = prob_P3L_if_node_fail.values
        min_value = min(prob_P3L_if_node_fail_values)
        max_value = max(prob_P3L_if_node_fail_values)
        min_value = round(min_value, 2) - 0.01  # Restrict minimum value to 2 decimal places
        max_value = round(max_value, 2) + 0.01  # Restrict maximum value to 2 decimal places and add 0.01
        normal_prob_P3L_if_node_fail = (prob_P3L_if_node_fail_values[0] - min_value) / (max_value - min_value)
        node_prob_dict[node] = normal_prob_P3L_if_node_fail
sorted_node_prob = sorted(node_prob_dict.items(), key=lambda x: x[1], reverse=True)

# Save results to Excel worksheet
#workbook = openpyxl.Workbook()
#worksheet = workbook.active
#worksheet['A1'] = 'Node'
#worksheet['B1'] = 'Probability'
#red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
#yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

#for index, (node, prob) in enumerate(sorted_node_prob, start=2):
#    print("index:", index, "node:", node, "prob:", prob)
#    cell_node = worksheet.cell(row=index, column=1)
#    cell_node.value = node    
#    cell_prob = worksheet.cell(row=index, column=2)
#    cell_prob.value = prob
    
#    if prob > 0.8:
#        cell_prob.fill = red_fill
#    elif prob >= 0.5 and prob <= 0.8:
#        cell_prob.fill = yellow_fill
#    else:
#        cell_prob.fill = green_fill
#print ("\n")
#workbook.save(filename='sorted_results.xlsx')

# Print results to console
for nodes in total_elements:
    if nodes==last_node:
        prob_termination=inference.query(variables=[nodes])
        print("\n")
        print("[*] CPT (Occurrence) of CPS System Termination:\n", prob_termination)
        impact_termination=inference2.query(variables=[nodes])
        print("[*] CPT (Severity/Impact) of CPS System Termination:\n", impact_termination)        
        cpd_prob = prob_termination.values
        cpd_impact = impact_termination.values
        print('--------------------------------------------------------')
        print("[*] Probability of CPS Termination:", cpd_prob[0])
        print("[*] Probability of Severity/Impact:", cpd_impact[0])
        risk_prob = cpd_prob[0]*cpd_impact[0]
        #print("Raw Risk Value:", risk_prob)
        normal_risk = (risk_prob) / (0.0674)
        print('[*] Current normalised risk score: {:.2f} %'.format(normal_risk * 100))
        print('--------------------------------------------------------')
        if normal_risk<0.2:
            print('[----] CPS System is under NEGLIGIBLE risk (less than 20%)')
        elif 0.2<=normal_risk<0.4:
            print('[*---] CPS System is under LOW risk (between 20% and 40%)')
        elif 0.4<=normal_risk<0.6:
            print('[**--] CPS System is under MEDIUM risk (between 40% and 60%)')
        elif 0.6<=normal_risk<0.8:
            print('[***-] CPS System is under HIGH risk (between 60% and 80%)')
        else:
            print('[****] CPS System is under CRITICAL risk (greater than 80%)')
    else:
        pass